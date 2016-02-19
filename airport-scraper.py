# airportScraper.py 
# Author: Lakshay Akula
# Date: 02/18/2016
# Description:
#  This simple python script scrapes Department of Transportation airport traffic data into
#  neat Excel spreadsheets. It can be adapted for other web scraping purposes as well. Read
#  more on https://github.com/Kommandat/AirportScraper/. 

# Run requests on the site
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

# Scrape HTML
import requests
from BeautifulSoup import BeautifulSoup

# Work with Excel
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

# Start Firefox browser
browser = webdriver.Firefox()
browser.get('http://www.transtats.bts.gov/Data_Elements.aspx?Data=1')

# Some predetermined airport hubs
los_angeles = ["- Los Angeles, CA: Los Angeles International",
			   "- Ontario, CA: Ontario International",
			   "- Santa Ana, CA: John Wayne Airport-Orange County",
			   "- Burbank, CA: Bob Hope",
			   "- Long Beach, CA: Long Beach Airport"]

bay = ["- San Francisco, CA: San Francisco International",
	   "- Oakland, CA: Metropolitan Oakland International",
	   "- San Jose, CA: Norman Y. Mineta San Jose International"]

boston = ["- Boston, MA: Logan International",
		  "- Manchester, NH: Manchester-Boston Regional",
	      "- Providence, RI: Theodore Francis Green State"]

chicago = ["- Chicago, IL: Chicago O\'Hare International",
		   "- Chicago, IL: Chicago Midway International",
		   "- Milwaukee, WI: General Mitchell International"]

houston = ["- Houston, TX: George Bush Intercontinental/Houston",
	       "- Houston, TX: William P Hobby"]

dallas = ["- Dallas/Fort Worth, TX: Dallas/Fort Worth International",
	      "- Dallas, TX: Dallas Love Field"]

new_york = ["- New York, NY: John F. Kennedy International",
			"- Newark, NJ: Newark Liberty International",
			"- New York, NY: LaGuardia"]

miami = ["- Fort Lauderdale, FL: Fort Lauderdale-Hollywood International",
		 "- Miami, FL: Miami International",
		 "- West Palm Beach/Palm Beach, FL: Palm Beach International"]

dc = ["- Washington, DC: Ronald Reagan Washington National",
	  "- Washington, DC: Washington Dulles International",
	  "- Baltimore, MD: Baltimore/Washington International Thurgood Marshall"]

cincinnati = ["- Cincinnati, OH: Cincinnati/Northern Kentucky International",
		      "- Dayton, OH: James M Cox/Dayton International",
		      "- Lexington, KY: Blue Grass"]

cleaveland = ["- Cleveland, OH: Cleveland-Hopkins International",
		   	  "- Akron, OH: Akron-Canton Regional"]

# All airport codes
airportCodes = {"All":"All",
 				"- Los Angeles, CA: Los Angeles International":"LAX",
				"- Ontario, CA: Ontario International":"ONT",
				"- Santa Ana, CA: John Wayne Airport-Orange County":"SNA",
				"- Burbank, CA: Bob Hope":"BUR",
				"- Long Beach, CA: Long Beach Airport":"LGB",

				"- San Francisco, CA: San Francisco International":"SFO",
				"- Oakland, CA: Metropolitan Oakland International":"OAK",
				"- San Jose, CA: Norman Y. Mineta San Jose International":"SJC",

				"- Boston, MA: Logan International":"BOS",
				"- Manchester, NH: Manchester-Boston Regional":"MHT",
				"- Providence, RI: Theodore Francis Green State":"PVD",

				"- Chicago, IL: Chicago O\'Hare International":"ORD",
				"- Chicago, IL: Chicago Midway International":"MDW",
				"- Milwaukee, WI: General Mitchell International":"MKE",

				"- Houston, TX: George Bush Intercontinental/Houston":"IAH",
				"- Houston, TX: William P Hobby":"HOU",

				"- Dallas/Fort Worth, TX: Dallas/Fort Worth International":"DFW",
				"- Dallas, TX: Dallas Love Field":"DAL",

				"- New York, NY: John F. Kennedy International":"JFK",
				"- Newark, NJ: Newark Liberty International":"EWR",
				"- New York, NY: LaGuardia":"LGA",

				"- Fort Lauderdale, FL: Fort Lauderdale-Hollywood International":"FLL",
				"- Miami, FL: Miami International":"MIA",
				"- West Palm Beach/Palm Beach, FL: Palm Beach International":"PBI",

				"- Washington, DC: Ronald Reagan Washington National":"DCA",
				"- Washington, DC: Washington Dulles International":"IAD",
				"- Baltimore, MD: Baltimore/Washington International Thurgood Marshall":"BWI",

				"- Cincinnati, OH: Cincinnati/Northern Kentucky International":"CVG",
				"- Dayton, OH: James M Cox/Dayton International":"DAY",
				"- Lexington, KY: Blue Grass":"LEX",

				"- Cleveland, OH: Cleveland-Hopkins International":"CLE",
				"- Akron, OH: Akron-Canton Regional":"CAK"}

# List of statistics you're interested in
stats       = ["Passengers", "Flights", "LoadFactor"]

# origin or dest
originOrDest = ["Origin", "Destination"]

# List of hubs you want to scrape
hubs     = [los_angeles, bay, boston, chicago, houston,
		        dallas, new_york, miami, dc, cincinnati, cleveland]

hubNames = ["Los Angeles", "Bay", "Boston", "Chicago", "Houston",
		        "Dallas", "New York", "Miami", "DC", "Cincinnati", "Cleveland"]

# Some predermined lists of carriers
allCarriers = ["All U.S. and Foreign Carriers",
      			   "All U.S. Carriers",
      			   "All Foreign Carriers",
      			   "Alaska Airlines",
      			   "American Airlines",
      			   "Atlas Air",
      			   "Delta Air Lines",
      			   "Envoy Air",
      			   "ExpressJet Airlines",
      			   "Frontier Airlines",
      			   "Hawaiian Airlines",
      			   "JetBlue Airways",
      			   "SkyWest Airlines",
      			   "Southwest Airlines",
      			   "Spirit Air Lines",
      			   "United Air Lines",
      			   "Virgin America"] 

# List of carriers you want to scrape
carriers = allCarriers

# We need this so we know which hubName to eventually make the file name
count = 0

for hub in hubs:

	# The Excel workbook we'll be outputting to
	wb = Workbook()

	for stat in stats:
		
		# pick our stat
		elem = browser.find_element_by_id("Link_" + stat)
		elem.click()
		
		for choice in originOrDest:

			# pick our choice of origin or dest
			elem = browser.find_element_by_id("Link_" + choice)
			elem.click()
      
      # Create a new sheet for each statistic and origin or destination
			ws = wb.create_sheet(title = stat + " " + choice)

			for airport in hub:

				for carrier in carriers:

					# select airport
					Select(browser.find_element_by_id("AirportList")).select_by_visible_text(airport)
					
					# select carrier
					Select(browser.find_element_by_id("CarrierList")).select_by_visible_text(carrier)
					
					# hit submit!
					elem = browser.find_element_by_id("Submit")
					elem.click()

					# now scrape the data
					html = browser.page_source
					
					soup = BeautifulSoup(html)
					table = soup.find('table', attrs={'class': 'dataTDRight'})

          # Only continue if the table is not None (i.e. there's data 
          # availible.
					if table is not None:

						list_of_rows = []
						for row in table.findAll('tr')[1:]:
						    list_of_cells = [carrier, airportCodes[airport]]
						    for cell in row.findAll('td'):
						        text = cell.text.replace('&nbsp;', '')
						        list_of_cells.append(text)
						    list_of_rows.append(list_of_cells)
						    
					  # Create the first row for the data
						ws.append(["Carrier", "Airport","Year", "Month","Domestic", "International", "Total"])

            # Add the rest of the data
						for i in range(len(list_of_rows)):
							ws.append(list_of_rows[i])
	
	# Create a new excel file for each hub
	wb.save(hubNames[count] + ".xlsx")
	count+=1

# Close firefox
browser.close()
